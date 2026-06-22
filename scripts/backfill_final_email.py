#!/usr/bin/env python3
"""
One-time backfill: set et_items.final_email_date from the Excel "F.Email Dt."
column, matched by title. Safe & idempotent — only touches final_email_date,
matches existing rows by id (looked up from title), and never deletes anything.

Run AFTER add_final_email.sql:
    python scripts/backfill_final_email.py "Work in Progress - 2026 (260622).xlsx"
"""
import sys
import os
import json
import datetime
import urllib.request
import urllib.error

import openpyxl

# F.Email column per sheet (1-indexed) + title column + first data row.
SHEETS = {
    "2026": {"title": 1, "final_email": 29, "data_start": 2},
    "Kanzul Madaris": {"title": 2, "final_email": 30, "data_start": 4},
}


def load_env(path=".env.local"):
    env = {}
    if not os.path.exists(path):
        sys.exit(f"ERROR: {path} not found. Run from the project root.")
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def iso(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (v.date() if isinstance(v, datetime.datetime) else v).isoformat()
    return None


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python scripts/backfill_final_email.py <workbook.xlsx>")
    path = sys.argv[1]
    if not os.path.exists(path):
        sys.exit(f"ERROR: workbook not found: {path}")

    env = load_env()
    base = env["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/") + "/rest/v1"
    key = env["SUPABASE_SERVICE_ROLE_KEY"]
    headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def req(method, pathq, body=None, prefer=None):
        h = dict(headers)
        if prefer:
            h["Prefer"] = prefer
        data = json.dumps(body).encode("utf-8") if body is not None else None
        r = urllib.request.Request(base + pathq, data=data, headers=h, method=method)
        try:
            with urllib.request.urlopen(r) as resp:
                raw = resp.read().decode("utf-8")
                return json.loads(raw) if raw else None
        except urllib.error.HTTPError as e:
            sys.exit(f"HTTP {e.code} on {method} {pathq}: {e.read().decode('utf-8')}")

    # Title -> final_email from the workbook.
    print(f"Reading {path} ...")
    wb = openpyxl.load_workbook(path, data_only=True)
    title_to_email = {}
    for name, cfg in SHEETS.items():
        ws = wb[name]
        for row in range(cfg["data_start"], ws.max_row + 1):
            title = clean(ws.cell(row, cfg["title"]).value)
            email = iso(ws.cell(row, cfg["final_email"]).value)
            if title and email:
                title_to_email[title] = email
    print(f"  {len(title_to_email)} items have a final email date in Excel.")

    # Fetch existing items (id + title), in pages to be safe.
    print("Fetching existing items from Supabase ...")
    items = []
    offset = 0
    while True:
        page = req("GET", f"/et_items?select=id,title&order=created_at.asc&limit=1000&offset={offset}")
        if not page:
            break
        items.extend(page)
        if len(page) < 1000:
            break
        offset += 1000
    by_title = {i["title"]: i["id"] for i in items}
    print(f"  {len(items)} items in DB.")

    updated, missing = 0, 0
    for title, email in title_to_email.items():
        item_id = by_title.get(title)
        if not item_id:
            missing += 1
            continue
        req("PATCH", f"/et_items?id=eq.{item_id}", body={"final_email_date": email}, prefer="return=minimal")
        updated += 1

    print(f"\nDONE. Updated {updated} items. {missing} Excel rows had no matching DB title.")


if __name__ == "__main__":
    main()

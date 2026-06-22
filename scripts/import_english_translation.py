#!/usr/bin/env python3
"""
One-time importer: Excel "Work in Progress" -> Supabase ET tables.

Reads the workbook and loads:
  - sheet "2026"            -> et_items (+ et_stages), board = main_2026 / magazine
  - sheet "Kanzul Madaris"  -> et_items (+ et_stages), board = kanzul_madaris
  - sheet "Workforce"       -> et_people

Current step / holder is NOT imported (the app computes it from the stage
dates). The Reminder / Magazine view-sheets are skipped (they are derived).

Usage (from project root, .env.local must contain NEXT_PUBLIC_SUPABASE_URL
and SUPABASE_SERVICE_ROLE_KEY):

    python scripts/import_english_translation.py "Work in Progress - 2026 (260622).xlsx"

The script DELETES all existing rows in et_stages / et_items / et_people
first, then re-imports, so it is safe to re-run.
"""
import sys
import os
import json
import uuid
import datetime
import urllib.request
import urllib.error

import openpyxl

# ---- Stage column maps: stage_code -> (name_col, sent_col, rcvd_col) 1-indexed ----
STAGE_ORDER = ["TR", "IF", "CM", "ED", "NR", "ST", "FF", "FPR"]
STAGE_SEQ = {code: i + 1 for i, code in enumerate(STAGE_ORDER)}

SHEET_2026 = {
    "header_row": 1,
    "data_start": 2,
    "title": 1, "type": 2, "rcv": 3, "words": 4,
    "stages": {
        "TR": (5, 6, 7), "IF": (8, 9, 10), "CM": (11, 12, 13), "ED": (14, 15, 16),
        "NR": (17, 18, 19), "ST": (20, 21, 22), "FF": (23, 24, 25), "FPR": (26, 27, 28),
    },
    "final_email": 29, "current_step": 30, "since": 32, "further": 33,
}

SHEET_KM = {
    "header_row": 3,
    "data_start": 4,
    "title": 2, "type": 3, "rcv": 4, "words": 5,
    "stages": {
        "TR": (6, 7, 8), "IF": (9, 10, 11), "CM": (12, 13, 14), "ED": (15, 16, 17),
        "NR": (18, 19, 20), "FF": (21, 22, 23), "FPR": (24, 25, 26), "ST": (27, 28, 29),
    },
    "final_email": 30, "current_step": None, "since": None, "further": None,
}


def load_env(path=".env.local"):
    env = {}
    if not os.path.exists(path):
        sys.exit(f"ERROR: {path} not found. Run from the project root.")
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def is_na(v):
    return isinstance(v, str) and v.strip().upper() == "NA"


def iso(v):
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    return None


def to_int(v):
    try:
        if v is None or (isinstance(v, str) and not v.strip()):
            return None
        return int(float(v))
    except (ValueError, TypeError):
        return None


def clean_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


class Supa:
    def __init__(self, url, key):
        self.base = url.rstrip("/") + "/rest/v1"
        self.key = key

    def _req(self, method, path, body=None, prefer=None):
        headers = {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
            "Content-Type": "application/json",
        }
        if prefer:
            headers["Prefer"] = prefer
        data = json.dumps(body).encode("utf-8") if body is not None else None
        req = urllib.request.Request(self.base + path, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read().decode("utf-8")
                return json.loads(raw) if raw else None
        except urllib.error.HTTPError as e:
            sys.exit(f"HTTP {e.code} on {method} {path}: {e.read().decode('utf-8')}")

    def delete_all(self, table):
        self._req("DELETE", f"/{table}?id=not.is.null", prefer="return=minimal")

    def insert(self, table, rows, chunk=200):
        for i in range(0, len(rows), chunk):
            self._req("POST", f"/{table}", body=rows[i:i + chunk], prefer="return=minimal")


def parse_item_sheet(ws, cfg, board_resolver):
    """Yield (item_dict, [stage_dicts]) for each valid data row."""
    max_col = ws.max_column
    for r in range(cfg["data_start"], ws.max_row + 1):
        title = clean_str(ws.cell(r, cfg["title"]).value)
        if not title:
            continue
        item_id = str(uuid.uuid4())
        typ = clean_str(ws.cell(r, cfg["type"]).value)
        board = board_resolver(typ)

        stages = []
        any_assigned = False
        all_received = True
        applicable_count = 0
        for code, (ncol, scol, rcol) in cfg["stages"].items():
            name = ws.cell(r, ncol).value
            sent = ws.cell(r, scol).value
            rcvd = ws.cell(r, rcol).value
            na = is_na(name)
            person = None if na else clean_str(name)
            sent_date = None if is_na(sent) else iso(sent)
            rcvd_date = None if is_na(rcvd) else iso(rcvd)
            stages.append({
                "item_id": item_id,
                "stage": code,
                "seq": STAGE_SEQ[code],
                "person": person,
                "sent_date": sent_date,
                "received_back_date": rcvd_date,
                "not_applicable": na,
            })
            if not na:
                applicable_count += 1
                if person or sent_date or rcvd_date:
                    any_assigned = True
                if not rcvd_date:
                    all_received = False

        # Derive status (explicit Excel "Current Step" wins when present).
        excel_step = clean_str(ws.cell(r, cfg["current_step"]).value) if cfg["current_step"] else None
        final_email = iso(ws.cell(r, cfg["final_email"]).value) if cfg["final_email"] else None
        status = None
        if excel_step:
            low = excel_step.lower()
            if low == "completed":
                status = "completed"
            elif low.startswith("pending"):
                status = "pending_assignment"
        if status is None:
            if applicable_count and all_received:
                status = "completed"
            elif final_email:
                status = "completed"
            elif not any_assigned:
                status = "pending_assignment"
            else:
                status = "in_progress"

        item = {
            "id": item_id,
            "title": title,
            "type": typ,
            "board": board,
            "received_date": iso(ws.cell(r, cfg["rcv"]).value),
            "word_count": to_int(ws.cell(r, cfg["words"]).value),
            "delivery_date": None,
            "priority": None,
            "status": status,
            "further_process": clean_str(ws.cell(r, cfg["further"]).value) if cfg["further"] else None,
        }
        yield item, stages


def parse_workforce(ws):
    people = []
    for r in range(2, ws.max_row + 1):
        name = clean_str(ws.cell(r, 1).value)
        skills = clean_str(ws.cell(r, 2).value)
        email = clean_str(ws.cell(r, 3).value)
        # Skip divider rows like "for books" / "for Quran" (name only, no skills/email).
        if not name:
            continue
        if not skills and not email and name.lower().startswith("for "):
            continue
        people.append({
            "id": str(uuid.uuid4()),
            "name": name,
            "skills": skills,
            "email": email,
            "working_hours": clean_str(ws.cell(r, 4).value),
            "dpr_link": clean_str(ws.cell(r, 5).value),
            "notes": clean_str(ws.cell(r, 6).value),
            "active": True,
        })
    return people


def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python scripts/import_english_translation.py <workbook.xlsx>")
    path = sys.argv[1]
    if not os.path.exists(path):
        sys.exit(f"ERROR: workbook not found: {path}")

    env = load_env()
    url = env.get("NEXT_PUBLIC_SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        sys.exit("ERROR: NEXT_PUBLIC_SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY missing in .env.local")

    print(f"Loading workbook: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)

    def board_2026(typ):
        return "magazine" if (typ or "").lower() == "mgz" else "main_2026"

    items, stages = [], []
    for it, st in parse_item_sheet(wb["2026"], SHEET_2026, board_2026):
        items.append(it)
        stages.extend(st)
    n_main = len(items)

    for it, st in parse_item_sheet(wb["Kanzul Madaris"], SHEET_KM, lambda t: "kanzul_madaris"):
        items.append(it)
        stages.extend(st)
    n_km = len(items) - n_main

    people = parse_workforce(wb["Workforce"])

    from collections import Counter
    status_counts = Counter(i["status"] for i in items)
    board_counts = Counter(i["board"] for i in items)

    print(f"  Parsed items:  {len(items)}  (2026={n_main}, kanzul_madaris={n_km})")
    print(f"    by board:  {dict(board_counts)}")
    print(f"    by status: {dict(status_counts)}")
    print(f"  Parsed stages: {len(stages)}")
    print(f"  Parsed people: {len(people)}")

    supa = Supa(url, key)
    print("Clearing existing ET tables...")
    supa.delete_all("et_stages")
    supa.delete_all("et_items")
    supa.delete_all("et_people")

    print("Inserting people...")
    supa.insert("et_people", people)
    print("Inserting items...")
    supa.insert("et_items", items)
    print("Inserting stages...")
    supa.insert("et_stages", stages)

    print("\nDONE. Imported %d items, %d stages, %d people." % (len(items), len(stages), len(people)))


if __name__ == "__main__":
    main()
